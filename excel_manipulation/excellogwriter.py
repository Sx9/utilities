import os, re, pathlib, time, sys
from datetime import datetime
# import xml.dom, xml.dom.minidom
# From https://www.datacamp.com/community/tutorials/python-xml-elementtree
# import xml.etree.ElementTree as ET

# Uses OpenPyXL to read the Excel File
# https://openpyxl.readthedocs.io/en/stable/
# Install it using [pip install openpyxl]

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, Fill, Border, PatternFill, Side, GradientFill, Alignment
# From https://stackoverflow.com/a/52494457/14923106
from openpyxl.styles import Alignment

# FOR MICROSOFT WORD FILE MANIPULATION
# Uses Python-DOCX to write the Word File
# https://python-docx.readthedocs.io/en/latest/
# Install it using [pip install python-docx]
# import docx
# from docx import Document
# from docx.shared import Inches
# From https://stackoverflow.com/questions/56658872/add-page-number-using-python-docx
# from docx.oxml import OxmlElement, ns
# From https://stackoverflow.com/questions/18595864/python-create-a-table-of-contents-with-python-docx-lxml
# from docx.oxml.ns import qn
# From https://stackoverflow.com/questions/57189055/why-cant-i-import-wd-align-paragraph-from-docx-enum-text
# from docx.enum.text import WD_PARAGRAPH_ALIGNMENT

class ExcelMLLogWriter(object):

    def __init__(self, strFileAbsolutePath):

        binSuccess = True
        self.strFileAbsolutePath = strFileAbsolutePath
        binSuccess = self.fnInitializer()
        self.binExists, strResults, objWorkbookFound, objTrainingWorksheetFound, objPredictionWorksheetFound = self.fnFindFile(strFileAbsolutePath)

        if self.intLogLevel >= 3:
            print("FILE SEARCH RESULTS: " + strResults)

        if self.binExists:
            # LOAD UP THE FILE TO UPDATE IT
            # WORKSHEET LOADED IN objWorksheet
            self.objWorkbook = objWorkbookFound
            self.objTrainingWorksheet = objTrainingWorksheetFound
            self.objPredictionWorksheet = objPredictionWorksheetFound

        elif not self.binExists and strResults == "FILE NOT FOUND":
            # FILE DOES NOT EXIST SO CREATE IT, PREPARE IT, LOAD IT
            self.objWorkbook, self.objTrainingWorksheet, self.objPredictionWorksheet = self.fnCreateFile(strFileAbsolutePath)

        else:
            raise ValueError("Workbook initialization failed. File [%s] exists but something wrong with it (%s). Will not overwrite it!" % (strFileAbsolutePath, strResults))

    def fnInitializer(self):

        # VARIABLE DECLARATION
        self.strVersion = "0030"
        self.intLogLevel = 3
        self.strAuthor = "sandipan.com"

        self.strRegex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
        self.strHyperLinkImage = "https://upload.wikimedia.org/wikipedia/commons/thumb/4/44/Icon_External_Link.svg/1200px-Icon_External_Link.svg.png"
        self.strHyperLinkImageName = "1200px-Icon_External_Link.svg.png"
        self.strSandipanImageName = "logo2-200x130.png"
        self.intTitlePageSpacers = 3

        # SET UP STYLE
        self.fntHeaderFont = Font(name = "Tahoma", size = 12, color = "444444", bold = True)
        self.sidSide = Side(border_style="double", color="444444")
        self.bdrBorder = Border(bottom = self.sidSide)
        self.filFill = PatternFill("solid", fgColor = "BBBBBB")
        self.fntEntryFont = Font(name = "Tahoma", size = 10, color = "444444", bold = False)

        self.intHeaderRow = 3
        self.strTrainWorksheetName = "TRAINING LOG"
        self.strPredictWorksheetName = "PREDICTION LOG"
        self.lstColumnID = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]

        self.lstTrainingColumns = ["RUN", "BIN/FIL", "MODEL ARCHITECTURE", "TRAINING TIMESTAMP", "EPOCHS", "MODEL FILE", "WEIGHTS FILE", "GRAPH FILE", "TRAINING ACCURACY"]
        self.lstTrainingColumnWidths = [15, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25]
        self.lstPredictionColumns = ["PREDICTION TIMESTAMP", "PREDICTION SET FILE", "ACCURACY", "F1 SCORE"]
        self.lstPredictionColumns = self.lstTrainingColumns + self.lstPredictionColumns
        self.lstPredictionColumnWidths = [15, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25]
        self.lstTrainingTimestampColumnName = "TRAINING TIMESTAMP"

        return True

    def fnFindFile(self, strFileAbsolutePath):

        # IF XLS FILE EXISTS
        if os.path.exists(strFileAbsolutePath):

            # IF IT OPENS
            try:
                objWorkbook = load_workbook(filename=strFileAbsolutePath)
            except:
                return False, "NOT EXCEL FILE", None, None, None

            # IF WORKSHEET IS FOUND
            lstWorksheets = objWorkbook.sheetnames

            #TRAINING
            if self.strTrainWorksheetName in lstWorksheets:
                objTrainingWorksheet = objWorkbook[self.strTrainWorksheetName]

                binColumnMatch = True
                for intCtr, strColumnName in enumerate(self.lstTrainingColumns):
                    # CHECK COLUMN HEADERS AND SEE IF THEY MATCH
                    if self.intLogLevel >= 4:
                        print(intCtr, strColumnName, "-", objTrainingWorksheet[self.lstColumnID[intCtr] + str(self.intHeaderRow)].value)
                    if objTrainingWorksheet[self.lstColumnID[intCtr] + str(self.intHeaderRow)].value == strColumnName:
                        pass
                    else:
                        binColumnMatch = False

                if not binColumnMatch:
                    return False, "TRAINING COLUMNS DO NOT MATCH", None, None, None

            else:
                return False, "TRAINING WORKSHEET NOT FOUND", None, None, None

            # PREDICTION
            if self.strPredictWorksheetName in lstWorksheets:
                objPredictionWorksheet = objWorkbook[self.strPredictWorksheetName]

                binColumnMatch = True
                for intCtr, strColumnName in enumerate(self.lstPredictionColumns):
                    # CHECK COLUMN HEADERS AND SEE IF THEY MATCH
                    if self.intLogLevel >= 4:
                        print(intCtr, strColumnName, "-", objPredictionWorksheet[self.lstColumnID[intCtr] + str(self.intHeaderRow)].value)
                    if objPredictionWorksheet[self.lstColumnID[intCtr] + str(self.intHeaderRow)].value == strColumnName:
                        pass
                    else:
                        binColumnMatch = False

                if not binColumnMatch:
                    return False, "PREDICTION COLUMNS DO NOT MATCH", None, None, None

            else:
                return False, "PREDICTION WORKSHEET NOT FOUND", None, None, None

        else:
                return False, "FILE NOT FOUND", None, None, None

        return True, "FILE OK", objWorkbook, objTrainingWorksheet, objPredictionWorksheet

    def fnCreateFile(self, strFileAbsolutePath):

        # CREATE AND PREPARE EXCEL FILE
        self.objWorkbook = Workbook()

        # TRAINING WORKSHEET
        self.objTrainingWorksheet = self.objWorkbook.active
        self.objTrainingWorksheet.title = self.strTrainWorksheetName
        self.objTrainingWorksheet.sheet_properties.tabColor = "8888FF"

        # CREATE THE HEADERS
        for intCtr, strHeader in enumerate(self.lstTrainingColumns):
            strCellReference = self.lstColumnID[intCtr] + str(self.intHeaderRow)
            self.objTrainingWorksheet[strCellReference] = strHeader
            objCell = self.objTrainingWorksheet[strCellReference]
            objCell.font = self.fntHeaderFont
            objCell.border = self.bdrBorder
            objCell.fill = self.filFill
            self.objTrainingWorksheet.column_dimensions[self.lstColumnID[intCtr]].width = self.lstTrainingColumnWidths[
                intCtr]

        # PREDICTION WORKSHEET
        self.objPredictionWorksheet = self.objWorkbook.create_sheet(self.strPredictWorksheetName)
        self.objPredictionWorksheet.sheet_properties.tabColor = "8888FF"

        # CREATE THE HEADERS
        for intCtr, strHeader in enumerate(self.lstPredictionColumns):
            strCellReference = self.lstColumnID[intCtr] + str(self.intHeaderRow)
            self.objPredictionWorksheet[strCellReference] = strHeader
            objCell = self.objPredictionWorksheet[strCellReference]
            objCell.font = self.fntHeaderFont
            objCell.border = self.bdrBorder
            objCell.fill = self.filFill
            self.objPredictionWorksheet.column_dimensions[self.lstColumnID[intCtr]].width = \
            self.lstPredictionColumnWidths[intCtr]

        # WRITE THE FILE
        self.objWorkbook.save(strFileAbsolutePath)
        if self.intLogLevel >= 3:
            print("Created file with %d + %d header columns" % (
            len(self.lstTrainingColumns), len(self.lstPredictionColumns)))

        return self.objWorkbook, self.objTrainingWorksheet, self.objPredictionWorksheet

    def fnTrainingUpdate(self, lstTrainingColumns):

        # FIND FIRST EMPTY ROW
        intRun = 0
        for intRow in range(self.intHeaderRow, 100000):
            if self.objTrainingWorksheet['A' + str(intRow)].value is None or self.objTrainingWorksheet['A' + str(intRow)].value == "":
                # HIT EMPTY ROW
                if self.intLogLevel >= 3:
                    print("FIRST EMPTY TRAINING ROW IS %d" % intRow)
                    break
            else:
                if intRow > self.intHeaderRow:
                    intRun = int(self.objTrainingWorksheet['A' + str(intRow)].value)

        self.objTrainingWorksheet["A" + str(intRow)] = intRun+1
        for intColCtr, strValue in enumerate(lstTrainingColumns):
            self.objTrainingWorksheet[self.lstColumnID[intColCtr+1] + str(intRow)] = strValue
        self.objWorkbook.save(self.strFileAbsolutePath)
        if self.intLogLevel >= 3:
            print("WROTE %d TRAINING COLUMNS AT ROW %d" % (len(lstTrainingColumns), intRow))

        return True

    def fnPredictionUpdate(self, strTrainingTimeStamp, lstPredictionColumns):

        # FIND WHICH COLUMN HAS TRAINING TIMESTAMP
        intTSColumn = 0
        for intCtr, strColumnName in enumerate(self.lstTrainingColumns):
            if self.lstTrainingTimestampColumnName == strColumnName:
                intTSColumn = intCtr
        strTSColumnID = self.lstColumnID[intTSColumn]

        # FIND THE TRAINING ROW DATA
        intRun = 0
        binFound = False
        for intRow in range(self.intHeaderRow+1, 100000):
            if self.objTrainingWorksheet[strTSColumnID + str(intRow)].value == strTrainingTimeStamp:
                binFound = True
                # RECORD ENTIRE ROW
                dicTrainingRecord = {}
                for intColCtr, strColumnName in enumerate(self.lstTrainingColumns):
                    dicTrainingRecord[strColumnName] = self.objTrainingWorksheet[self.lstColumnID[intColCtr] + str(intRow)].value
                if self.intLogLevel >= 3:
                    print("FOUND TRAINING RECORD: [%s]" % dicTrainingRecord)
                break
        if not binFound:
            raise ValueError("TRAINING TIMESTEMP [%s] NOT FOUND IN FILE [%s] TRAINING LOG" % (strTrainingTimeStamp, self.strFileAbsolutePath))

        # FIND FIRST EMPTY ROW IN PREDICTION LOG SHEET
        intRun = 0
        for intRow in range(self.intHeaderRow, 100000):
            if self.objPredictionWorksheet['A' + str(intRow)].value is None or self.objPredictionWorksheet['A' + str(intRow)].value == "":
                # HIT EMPTY ROW
                if self.intLogLevel >= 3:
                    print("FIRST EMPTY PREDICTION ROW IS %d" % intRow)
                    break
            else:
                if intRow > self.intHeaderRow:
                    intRun = int(self.objPredictionWorksheet['A' + str(intRow)].value)

        self.objPredictionWorksheet["A" + str(intRow)] = intRun+1
        for intColCtr, strValue in enumerate(self.lstPredictionColumns):
            # FOR OTHER COLUMNS
            if intColCtr > 0:
                # FOR TRAINING COLUMNS, LOAD FROM DICTIONARY
                if intColCtr < len(self.lstTrainingColumns):
                    self.objPredictionWorksheet[self.lstColumnID[intColCtr] + str(intRow)] = dicTrainingRecord.get(strValue)
                # FOR PREDICTION COLUMNS, LOAD FROM LIST PROVIDED
                else:
                    self.objPredictionWorksheet[self.lstColumnID[intColCtr] + str(intRow)] = lstPredictionColumns[intColCtr - len(self.lstTrainingColumns)]
        self.objWorkbook.save(self.strFileAbsolutePath)
        if self.intLogLevel >= 3:
            print("WROTE %d PREDICTION COLUMNS AT ROW %d" % (len(lstPredictionColumns), intRow))

        return True
